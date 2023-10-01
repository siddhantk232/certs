from xml.dom.minidom import parseString
from openpyxl import load_workbook
from cairosvg import svg2pdf

wb = load_workbook("./response.xlsx")
ws = wb.active  # first worksheet


def getCol(name):
    return map(lambda x: x.value, filter(lambda x: x.value != None, ws[name]))


data = list(zip(getCol("C"), getCol("D")))  # name and email

head = data[0]


def getText(nodeList):
    res = []
    for node in nodeList:
        if node.nodeType == node.TEXT_NODE:
            res.append(node.data)
    return "".join(res)


def apply(text):
    res = []
    for cell in data[1:]:
        imi = text
        for hd in enumerate(head):
            imi = imi.replace(f"{{{hd[1]}}}", cell[hd[0]])
        res.append((cell[0], imi))
    return res


with open("./cert.svg") as cert:
    string = cert.read()
    res = apply(string)
    for output, svg in res:
        svg = arr = bytes(svg, "utf-8")
        svg2pdf(
            bytestring=svg,
            write_to=f"{output}.pdf",
            scale=1,
            output_width=640 * 2,
            output_height=480 * 2,
            parent_width=640,
            parent_height=480,
        )
